import vk_api
from vk_api import VkUpload
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
import requests
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt
import openpyxl
import re
import datetime
from pprint import pprint
import PIL.Image as Image

d = print

vk_session = vk_api.VkApi(
    token="################SECRET_INFORMATION##########################")
vk = vk_session.get_api()
longpoll = VkLongPoll(vk_session)

CMD_ERROR_MESSAGE = 'Вы не ввели свою группу.' \
                    '\nПожалуйста, введи свою группу, чтобы я мог ее запомнить'
CMD_HELP_MESSAGE = 'Я Запомнил твою группу ' \
                   '\nОсновной список комманд: ' \
                   '\nНапиши "бот" для вызова комманд.' \
                   '\n Бот "день недели" - узнать на день недели." ' \
                   '\nБот "Название группы" - узнать расписание другой группы ' \
                   '\n"Бот погода" - Показывает сегодняшнюю погоду' \
                   '\nКорона- чтобы узнать информацию по поводу Короновируса в России' \
                   '\nПогода - чтобы открыть меню погоды' \
                   '\nКорона "область" - чтобы узнать состояние коронавируса в области'

CORONA = '{}\n\nРегион: {}' \
         '\nСлучаев:{} ({} за сегодня)' \
         '\nАктивных:{}({} за сегодня)' \
         '\nВылечено:{} ({} за сегодня)' \
         '\nУмерло: {} ({} за сегодня)'

weekDate = datetime.timedelta(weeks=1)
nowDate = datetime.datetime.now().date()
beginStudy = datetime.date(2022, 2, 7)
howWeek = (nowDate - beginStudy) / weekDate
fileSchedule = []
groups = {}
students = []
month_l = ["января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа", "сентября", "октября", "ноября",
           "декабря"]
days_d = {"понедельник": 0, "вторник": 1, "среда": 2, "четверг": 3, "пятница": 4, "суббота": 5}
professor_name = ""

page = requests.get("https://www.mirea.ru/schedule/")
soup = BeautifulSoup(page.text, "html.parser")
result = soup.find("div", {"class": "rasspisanie"}). \
    find(string="Институт информационных технологий"). \
    find_parent("div"). \
    find_parent("div"). \
    findAll("a", class_="uk-link-toggle")  # получить ссылки
result.pop(0)
result.pop(0)
result.pop(0)

for x in result:
    f = open(f"{result.index(x)}.xlsx", "wb")  # открываем файл для записи, в режиме wb
    resp = requests.get(x["href"])  # запрос по ссылке
    f.write(resp.content)

month_l = ["января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа", "сентября", "октября", "ноября",
           "декабря"]
days_d = {"понедельник": 0, "вторник": 1, "среда": 2, "четверг": 3, "пятница": 4, "суббота": 5}


# row_index=2
# col_index=1
# book = openpyxl.load_workbook("file.xlsx") # открытие файла
# sheet = book.active # активный лист
# num_cols = sheet.max_column # количество столбцов
# num_rows = sheet.max_row # количество строк
# print(sheet.cell(row = row_index, column = col_index).value) # ячейка


def keyboard_1(event, vk):
    keyboard = VkKeyboard(one_time=True)
    keyboard.add_button('На сегодня', color=VkKeyboardColor.POSITIVE)
    keyboard.add_button('На завтра', color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button('На эту неделю', color=VkKeyboardColor.PRIMARY)
    keyboard.add_button('На следующую неделю', color=VkKeyboardColor.PRIMARY)
    keyboard.add_line()
    keyboard.add_button('Какая неделя?', color=VkKeyboardColor.SECONDARY)
    keyboard.add_button('Какая группа?', color=VkKeyboardColor.SECONDARY)

    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        keyboard=keyboard.get_keyboard(),
        message='Показать расписание: '
    )


def keyboard_4(event, vk):
    keyboard = VkKeyboard(one_time=True)
    keyboard.add_button("Сейчас", color=VkKeyboardColor.PRIMARY)
    keyboard.add_button("Сегодня", color=VkKeyboardColor.POSITIVE)
    keyboard.add_button("Завтра", color=VkKeyboardColor.POSITIVE)
    keyboard.add_line()
    keyboard.add_button("На 5 дней", color=VkKeyboardColor.POSITIVE)

    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        keyboard=keyboard.get_keyboard(),
        message=f"Показать погоду в Москве"
    )


def keyboard_2(event, vk, professor):
    keyboard = VkKeyboard(one_time=True)
    for i in professor:
        keyboard.add_button(i, color=VkKeyboardColor.PRIMARY)
    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        keyboard=keyboard.get_keyboard(),
        message="Выбери одного?"
    )


# Клавиатура расписания преподавателя
def keyboard_3(event, vk, professor_name):
    keyboard = VkKeyboard(one_time=True)
    keyboard.add_button("На сегодня", color=VkKeyboardColor.POSITIVE)
    keyboard.add_button("На завтра", color=VkKeyboardColor.NEGATIVE)
    keyboard.add_line()
    keyboard.add_button("На эту неделю", color=VkKeyboardColor.PRIMARY)
    keyboard.add_button("На следующую неделю", color=VkKeyboardColor.PRIMARY)

    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        keyboard=keyboard.get_keyboard(),
        message=f"Показать расписание преподавателя {professor_name}"
    )


# Проверка на наличие группы у пользователя
def group_ident(event, vk):
    if event.user_id in groups:
        return (True)
    else:
        send_message(event, vk, "Введи ГРУППУ ПОЖАЛУЙСТА!!!")
        return (False)


def id_ident(user_id):
    for i in students:
        if i == user_id:
            return True
    return False


def send_message(event, vk, msg):
    vk.messages.send(
        user_id=event.user_id,
        random_id=get_random_id(),
        message=msg
    )


def send_message_with_attachments(user_id, vk, msg, attachments):
    vk.messages.send(
        user_id=user_id,
        attachment=','.join(attachments),
        random_id=get_random_id(),
        message=msg
    )


def getSchedule(userGroup, date):
    print("Запустил поиск расписания группы")
    even = 1 - (int((date - beginStudy) / weekDate) + 1) % 2
    day = date.weekday()
    dayList = day * 12 + 4
    cell = "Расписание на {} {}:\n".format(date.day, month_l[date.month - 1])
    if userGroup[9] == "1":
        book = openpyxl.load_workbook("0.xlsx")
        sheet = book.active
    elif userGroup[9] == "0":
        book = openpyxl.load_workbook("1.xlsx")
        sheet = book.active
    elif userGroup[9] == "9":
        book = openpyxl.load_workbook("2.xlsx")
        sheet = book.active
    else:
        return "Я не смог найти расписание для этой группы"
    num_cols = sheet.max_column  # количество столбцов
    num_rows = sheet.max_row  # количество строк
    for i in range(1, num_cols):
        if sheet.cell(row=2, column=i).value == userGroup:
            for j in range(dayList + even, dayList + 12, 2):
                if sheet.cell(row=j, column=i).value:
                    cell += "{}) {}, {}, {}, {}\n".format((j - dayList) // 2 + 1,
                                                          sheet.cell(row=j, column=i).value,
                                                          sheet.cell(row=j, column=i + 1).value,
                                                          sheet.cell(row=j, column=i + 2).value,
                                                          sheet.cell(row=j, column=i + 3).value)
                else:
                    cell += "{}) —\n".format((j - dayList) // 2 + 1)
    return cell.replace("None", "—")


def wind_ident_1(wind):
    if wind < 0.2:
        return ("штиль")
    elif wind >= 0.2 and wind < 1.5:
        return ("ихий")
    elif wind >= 1.5 and wind < 3.3:
        return ("легкий")
    elif wind >= 3.3 and wind < 5.4:
        return ("слабый")
    elif wind >= 5.4 and wind < 7.9:
        return ("умеренный")
    elif wind >= 7.9 and wind < 10.7:
        return ("свежий")
    elif wind >= 10.7 and wind < 13.8:
        return ("сильный")
    elif wind >= 13.8 and wind < 17.1:
        return ("крепкий")
    elif wind >= 17.1 and wind < 20.7:
        return ("очень крепкий")
    elif wind >= 20.7 and wind < 24.4:
        return ("шторм")
    elif wind >= 24.4 and wind < 28.4:
        return ("сильный шторм")
    elif wind >= 28.4 and wind < 32.6:
        return ("жестокий шторм")
    elif wind >= 32.6:
        return ("ураган")


# Направление ветра в погоде (для getWeather)
def wind_ident_2(wind):
    if wind >= 337.5 or wind < 22.5:
        return ("северный")
    elif wind >= 22.5 and wind < 67.5:
        return ("северо-восточный")
    elif wind >= 67.5 and wind < 112.5:
        return ("восточный")
    elif wind >= 112.5 and wind < 157.5:
        return ("юго-восточный")
    elif wind >= 157.5 and wind < 202.5:
        return ("южный")
    elif wind >= 202.5 and wind < 247.5:
        return ("юго-западный")
    elif wind >= 247.5 and wind < 292.5:
        return ("западный")
    elif wind >= 292.5 and wind < 337.5:
        return ("северо-западный")


# Словарик погоды (description)
def getDescription(description):
    name_1 = {"clear sky": "Ясное небо", "few clouds": "Местами облачно",
              "scattered clouds": "Рассеянная облачность",
              "broken clouds": "Облачность с просветами",
              "overcast clouds": "Пасмурная облачность",
              "shower rain": "Мелкий дождь", "rain": "Дождь",
              "thunderstorm": "Гроза", "snow": "Снег", "mist": "Туман",
              "tornado": "Торнадо", "squalls": "Шквалы",
              "volcanic ash": "Вулканический пепел", "dust": "Пыль",
              "fog": "Туман", "Heavy shower snow": "Сильный снегопад",
              "sand/ dust whirls": "Вихри песка/пыли", "sand": "Песок",
              "Haze": "Дымка", "Light shower snow": "Легкий снегопад",
              "Smoke": "Смог", "Shower sleet": "Дождь с мокрым снегом",
              "Shower snow": "Снегопад", "light snow": "Небольшой снег",
              "Rain and snow": "Снег с дождем", "Sleet": "Мокрый снег",
              "Light rain and snow": "Снег с небольшим дождем",
              "Light shower sleet": "Небольшой дождь с мокрым снегом",
              "Heavy snow": "Сильный снег", "shower rain": "Ливень",
              "ragged shower rain": "Неровный дождь с дождем",
              "heavy intensity shower rain": "Ливень с сильной интенсивностью",
              "light intensity shower rain": "Ливень небольшой интенсивности",
              "freezing rain": "Ледяной дождь", "extreme rain": "Сильный дождь",
              "very heavy rain": "Очень сильный дождь",
              "heavy intensity rain": "Дождь сильной интенсивности",
              "moderate rain": "Умеренный дождь", "light rain": "Небольшой дождь",
              "shower drizzle": "Моросящий дождь",
              "heavy shower rain and drizzle": "Сильный ливень и морось",
              "shower rain and drizzle": "Ливень с дождем и моросью",
              "heavy intensity drizzle rain": "Сильный моросящий дождь",
              "drizzle rain": "Моросящий дождь", "drizzle": "Изморось",
              "light intensity drizzle rain": "Моросящий дождь небольшой интенсивности",
              "heavy intensity drizzle": "Сильный моросящий дождь",
              "light intensity drizzle": "Изморось легкой интенсивности",
              "thunderstorm with light rain": "Гроза с небольшим дождем",
              "thunderstorm with rain": "Гроза с дождем",
              "thunderstorm with heavy rain": "Гроза с сильным дождем",
              "light thunderstorm": "Легкая гроза",
              "heavy thunderstorm": "Сильная гроза",
              "ragged thunderstorm": "Гроза с просветами",
              "thunderstorm with light drizzle": "Гроза с легким моросящий дождь",
              "thunderstorm with drizzle": "Гроза с моросящим дождем",
              "thunderstorm with heavy drizzle": "Гроза с сильным моросящим дождем",
              }
    if description in name_1:
        return (name_1[description])
    else:
        return (description)


# Словарик погоды (main)
def getMain(main):
    name_2 = {"Thunderstorm": "гроза", "Drizzle": "изморось",
              "Rain": "дождь", "Snow": "снег", "Mist": "тумам",
              "Smoke": "смог", "Haze": "дым", "Dust": "пыль",
              "Fog": "туман", "Sand": "песок", "Ash": "пепел",
              "Squall": "шквал", "Tornado": "торнадо",
              "Clear": "ясно", "Clouds": "облачно"}
    if main in name_2:
        return (name_2[main])
    else:
        return (main)


def getWeather(n):
    token = "70f6144a2107a75fa96a078fbb4a4660"
    r = requests.get(f"http://api.openweathermap.org/data/2.5/weather?q=moscow&appid={token}&units=metric")
    data = r.json()
    weathermain = data["weather"][0]["main"]
    weather = data["weather"][0]["description"]
    weatherid = data["weather"][0]["id"]
    tempmin = data["main"]["temp_min"]
    tempmax = data["main"]["temp_max"]
    comb = str(tempmin) + "-" + str(tempmax)
    winds = data["wind"]["speed"]
    windd = data["wind"]["deg"]
    pressure = data["main"]["pressure"]
    humidity = data["main"]["humidity"]
    icon = data["weather"][0]["icon"]
    if n == 1:
        s = (
            f"{getDescription(weather)}\nТемпература:{comb}°С\nДавление: {pressure} мм рт. ст., влажность:{humidity}%\nВетер: {wind_ident_1(winds)}, {winds} м/с, {wind_ident_2(windd)}")
        return (s)
    elif n == 2:
        return (f"http://openweathermap.org/img/wn/{icon}@2x.png")


def getWeater_day(n):
    token = "70f6144a2107a75fa96a078fbb4a4660"
    r = requests.get(f"http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid={token}&units=metric")
    data = r.json()
    temp = ""
    info = ""
    day = ["УТРО", "ДЕНЬ", "ВЕЧЕР", "НОЧЬ"]
    if n == 1:
        a = 0
        b = 4
    if n == 2:
        a = 4
        b = 8
    for i in range(a, b):
        temp += "/" + str(data["list"][i]["main"]["temp"]) + "/"
        info += f'{day[i - a]}\n'
        info += f'//{getDescription(data["list"][i]["weather"][0]["description"])}, температура: {str(data["list"][i]["main"]["temp_min"])} - {str(data["list"][i]["main"]["temp_max"])}°С\n'
        info += f'//Давление: {str(data["list"][i]["main"]["pressure"])} мм рт. ст., влажность{str(data["list"][i]["main"]["humidity"])}%\n'
        info += f'//Ветер: {wind_ident_1(data["list"][i]["wind"]["speed"])}, {str(data["list"][i]["wind"]["speed"])} м/с, {wind_ident_2(data["list"][i]["wind"]["deg"])}\n'
    temp += "\n\n"
    return (temp + info)


def getWeater_day5():
    token = "70f6144a2107a75fa96a078fbb4a4660"
    r = requests.get(f"http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid={token}&units=metric")
    data = r.json()
    temp_1 = ""
    temp_2 = ""
    for i in range(1, 20, 4):
        temp_1 += "/" + str(data["list"][i]["main"]["temp"]) + "/"
    for i in range(2, 20, 4):
        temp_2 += "/" + str(data["list"][i]["main"]["temp"]) + "/"
    temp_1 += "ДЕНЬ\n"
    temp_2 += "НОЧЬ"
    return (temp_1 + temp_2)


# Ищем нужную картинку
def getPic(n):
    token = "70f6144a2107a75fa96a078fbb4a4660"
    r = requests.get(f"http://api.openweathermap.org/data/2.5/forecast?q=moscow&appid={token}&units=metric")
    data = r.json()
    if n == 1:
        a = 0
        b = 4
        c = 1
    if n == 2:
        a = 4
        b = 8
        c = 1
    if n == 3:
        a = 1
        b = 20
        c = 4
    for i in range(a, b, c):
        icon = data["list"][i]["weather"][0]["icon"]
        image = requests.get(f"http://openweathermap.org/img/wn/{icon}@2x.png", stream=True)
        with open(f"file{i}.png", "wb") as f:
            f.write(image.content)
    if n == 1 or n == 2:
        img = Image.new('RGBA', (400, 100))
    elif n == 3:
        img = Image.new('RGBA', (500, 100))
    img0 = Image.open("file0.png")
    img1 = Image.open("file1.png")
    img2 = Image.open("file2.png")
    img3 = Image.open("file3.png")
    img.paste(img0, (0, 0))
    img.paste(img1, (100, 0))
    img.paste(img2, (200, 0))
    img.paste(img3, (300, 0))
    if n == 3:
        img4 = Image.open("file4.png")
        img.paste(img4, (400, 0))
    img.save("image.png")


def professorSearch(event, vk, name):
    print("Запустил поиск преподавателя")
    professor = []
    for n in range(3):
        book = openpyxl.load_workbook(f"{n}.xlsx")
        print(f"Открыл файл {n}")
        sheet = book.active
        num_cols = sheet.max_column  # количество столбцов
        num_rows = sheet.max_row  # количество строк
        for i in range(1, num_cols):
            if sheet.cell(row=3, column=i).value == "ФИО преподавателя":
                for j in range(4, num_rows):
                    if isinstance(sheet.cell(row=j, column=i).value, str):
                        if re.findall(name, sheet.cell(row=j, column=i).value):
                            result = re.findall(r"\w+", sheet.cell(row=j, column=i).value)
                            for i in range(len(result)):
                                if result[i] == name:
                                    name = result[i] + " " + result[i + 1] + "." + result[i + 2] + "."
                                    if not name in professor:
                                        professor.append(name)
    return (professor)


# Получение расписания препода
def getProfessor(professor_name, date):
    print("Запустил поиск расписания преподавателя")
    even = 1 - (int((date - beginStudy) / weekDate) + 1) % 2  # Четность недели
    day = date.weekday()  # День недели
    dayList = day * 12 + 4  # Номер строки на данный день недели (например понедельник)
    cell = []
    cell_2 = ["1) -\n", "2) -\n", "3) -\n", "4) -\n", "5) -\n", "6) -\n"]  # Тот самый костыль
    cell_3 = "Расписание на {} {}:\n".format(date.day, month_l[date.month - 1])
    # Проходимся по 3 файлам
    for n in range(3):
        book = openpyxl.load_workbook(f"{n}.xlsx")
        sheet = book.active
        num_cols = sheet.max_column  # Столбцы
        num_rows = sheet.max_row  # Строки
        # Цикл со строками в промежутке от начала дня недели до конца (по строкам понедельника идет)
        for i in range(dayList + even, dayList + 12, 2):
            # Ну тут по всем столбцам, если есть фамилия препода, то записываем предмет, номер группы, аудиторию..
            for j in range(1, num_cols):
                if sheet.cell(row=3, column=j).value == "ФИО преподавателя":
                    if re.findall(professor_name, str(sheet.cell(row=i, column=j).value)):
                        if sheet.cell(row=i, column=j - 2).value:
                            cell.append("{}) {}, {}, {}, {}\n".format(sheet.cell(row=i, column=2).value,
                                                                      sheet.cell(row=i, column=j - 2).value,
                                                                      sheet.cell(row=i, column=j - 1).value,
                                                                      sheet.cell(row=i, column=j + 1).value,
                                                                      sheet.cell(row=2, column=j - 2).value))
    # Подставляем то что получилось в костыль)
    for i in cell:
        cell_2[int(i[0]) - 1] = i
    cell_3 += "".join(cell_2)
    return (cell_3)


# Возврат короны за последние 10 дней
def get_corona_all_stat():
    try:
        page = requests.get('https://coronavirusstat.ru/country/russia/')  # Получаем страницу
        soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
        result = soup.find('table', {'class': 'table table-bordered small'}).findAll('tr')

        days = []  # День
        active = []  # Выздоровели сейчас
        cured = []  # Болеют
        died = []  # Умерли
        cases = []
        stats = []
        ml = 1000000
        print(result)
        for i in range(1, 11):  # цикл 10 раз
            days.append(result[i].find('th').getText())
            for a in result[i].findAll('td'):
                stats.append(int(a.getText().split(' ')[1]))
        for i in range(0, len(stats), 4):
            active.append(stats[i] / ml)
        for i in range(1, len(stats), 4):
            cured.append(stats[i] / ml)
        for i in range(2, len(stats), 4):
            died.append(stats[i] / ml)
        for i in range(3, len(stats), 4):
            cases.append(stats[i] / ml)
        # Нужно перевернуть чтобы даты были не 30,29,28 а 28,29,30. Также и с отсальными параметрами
        days = list(reversed(days))
        active = list(reversed(active))
        cured = list(reversed(cured))
        died = list(reversed(died))
        cases = list(reversed(cases))
    except:
        days = active = cured = died = cases = [-1] * 4

    return days, active, cured, died, cases


# Получение статистики короны, как для региона, так и для России
def get_corona_stat(special_url):
    # Статистика за сегодняшний день
    page = requests.get('https://coronavirusstat.ru' + special_url)  # Получаем страницу
    soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
    result = soup.find(string='Прогноз заражения на 10 дней').find_parent('div', {
        'class': 'border rounded mt-3 mb-3 p-3'})
    status = result.find('h6', 'text-muted').getText()[:-17]
    data = result.findAll('div', {'class': 'col col-6 col-md-3 pt-4'})
    plus = [] * 4
    value = [] * 4
    for i in range(4):
        value.append(data[i].find('div', 'h2').getText())
        plus.append(data[i].find('span', {'class': 'font-weight-bold'}).getText())
    return status, value, plus


# Показывает статискику и график Короны для России
def show_corona_all_stat(user_id):
    url = ''  # Это так называемый костыль, чтобы функция с получением данных не сломалась
    days, active, cured, died, cases = get_corona_all_stat()
    graf_data = {
        'Активных': active,
        'Вылечено': cured,
        'Умерло': died,
    }
    for i in range(len(days)):
        days[i] = days[i][:-5]
    fig, ax = plt.subplots()
    ax.stackplot(days, graf_data.values(), labels=graf_data.keys(), alpha=0.8)
    ax.legend(loc='upper left')
    ax.set_title('Коронавирус статистика по России')  # Делаем Верхний титульник
    ax.set_ylabel('Количество - Миллионы')  # Делаем боковой титульник
    fig.savefig("graf.png")  # Сохранение пнг
    attachments = []
    upload = VkUpload(vk_session)
    photo = upload.photo_messages("graf.png")[0]
    attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
    send_message_with_attachments(user_id, vk, reform_corona('Россия', get_corona_stat(url)), attachments)


# Показывает статистику для области
def show_corona_region_stat(user_id, region_list):
    region = region_list.split(' ')
    print(region[1])
    page = requests.get('https://coronavirusstat.ru/country/russia')  # Получаем страницу
    soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
    result = soup.findAll('div',
                          {'class': 'c_search_row'})  # Все блоки div со значением класса c_search_row он записывает
    d = ''
    rg = 'Россия'
    for x in result:  # Прогоняем все нужные блоки
        tmp = x.find('span', 'small').find('a')
        # print(tmp)
        if region[1] in tmp.getText().split(
                ' '):  # split нужен чтобы вычеркнуть слово область к примеру и осталось только Мурмонская
            rg = tmp.getText()
            d = tmp.get('href')  # href это ссылка на определенную область заданной в блоке div
            break
    send_message(user_id, vk, reform_corona(rg, get_corona_stat(d)))  # Вывод региона


# Функция чтобы все выглядило красиво
def reform_corona(region, data):
    status, value, plus = data
    return CORONA.format(status, region, value[0], plus[0], value[1], plus[1], value[2], plus[2], value[3],
                         plus[3])  # Корона написана выше, это так называемая константа для шаблона


def main():
    professor = []
    upload = VkUpload(vk_session)
    for event in longpoll.listen():
        if event.type == VkEventType.MESSAGE_NEW and event.text and event.to_me:
            if id_ident(event.user_id) == False:
                keyboardd = VkKeyboard(one_time=True)
                keyboardd.add_button('Привет', color=VkKeyboardColor.POSITIVE)
                students.append(event.user_id)
            print("Новая заявка {}, сообщение = {}".format(event.user_id, event.text))
            # Комманда ПРИВЕТ или НАЧАТЬ
            if event.text.lower() == "привет" or event.text.lower() == "начать":
                msg = "Приветствую, " + vk.users.get(user_id=event.user_id)[0][
                    "first_name"] + ". Я умный бот, который поможет тебе узнать расписание\nДля продолжения работы, введите свою группу:"
                send_message(event, vk, msg)

                # Комманда НОМЕР ГРУППЫ
            elif re.fullmatch(r"\w{4}-\d{2}-\d{2}", event.text):
                groups[event.user_id] = event.text.upper()
                send_message(event, vk, CMD_HELP_MESSAGE)

                # Комманда БОТ
            elif event.text.lower() == "бот":
                professor = []
                if event.user_id in groups:
                    keyboard_1(event, vk)
                else:
                    send_message(event, vk, CMD_ERROR_MESSAGE)

                # Комманда КАКАЯ НЕДЕЛЯ
            elif event.text.lower() == "какая неделя?":
                msg = 'Идет ' + str(int(howWeek) + 1) + " учебная неделя"
                send_message(event, vk, msg)

                # Комманда КАКАЯ ГРУППА
            elif event.text.lower() == "какая группа?":
                if group_ident(event, vk):
                    gotRec = groups[event.user_id]
                    msg = 'Расписание для группы ' + gotRec
                    send_message(event, vk, msg)
                else:
                    send_message(event, vk, CMD_ERROR_MESSAGE)

                # Комманда НА СЕГОДНЯ
            elif event.text.lower() == "на сегодня" and not professor:
                if group_ident(event, vk):
                    if nowDate.weekday() != 7:
                        gotRec = groups[event.user_id]
                        msg = str(getSchedule(gotRec, nowDate))
                        send_message(event, vk, msg)
                    else:
                        msg = "Сегодня выходной"
                        send_message(event, vk, msg)

                # Комманда НА ЗАВТРА
            elif event.text.lower() == "на завтра" and not professor:
                if group_ident(event, vk):
                    if (nowDate + datetime.timedelta(days=1)).weekday() != 7:
                        gotRec = groups[event.user_id]
                        msg = str(getSchedule(gotRec, nowDate + datetime.timedelta(days=1)))
                        send_message(event, vk, msg)
                    else:
                        msg = "Завтра выходной"
                        send_message(event, vk, msg)

                # Комманда НА ЭТУ НЕДЕЛЮ
            elif event.text.lower() == "на эту неделю" and not professor:
                if group_ident(event, vk):
                    gotRec = groups[event.user_id]
                    weekSchedule = []
                    for i in range(6):
                        weekSchedule.append(
                            getSchedule(gotRec, nowDate - datetime.timedelta(days=nowDate.weekday() - i)) + "\n\n")
                    msg = weekSchedule
                    send_message(event, vk, msg)

                # Комманда НА СЛЕД НЕДЕЛЮ
            elif event.text.lower() == "на следующую неделю" and not professor:
                if group_ident(event, vk):
                    gotRec = groups[event.user_id]
                    weekSchedule = []
                    for i in range(6):
                        weekSchedule.append(getSchedule(gotRec,
                                                        nowDate + datetime.timedelta(weeks=1) - datetime.timedelta(
                                                            days=nowDate.weekday() - i)) + "\n\n")
                    msg = weekSchedule
                    send_message(event, vk, msg)

                    # Комманда БОТ ДЕНЬ НЕДЕЛИ
            elif re.fullmatch(r"бот (понедельник|вторник|среда|четверг|пятница|суббота)", event.text.lower()):
                if group_ident(event, vk):
                    gotRec = groups[event.user_id]
                    msg = "Нечетная неделя: \n{} \n Четная неделя:\n{} ".format(
                        getSchedule(gotRec, beginStudy + datetime.timedelta(days=days_d[event.text[4:]]))[25:],
                        getSchedule(gotRec, beginStudy + weekDate + datetime.timedelta(days=days_d[event.text[4:]]))[
                        25:])
                    send_message(event, vk, msg)

            elif re.fullmatch(r"бот (понедельник|вторник|среда|четверг|пятница|суббота) и.бо-\d\d-\d\d",
                              event.text.lower()):
                gotText = event.text[4:].split()
                msg = "Нечетная неделя: {} \n Четная неделя:{} ".format(
                    getSchedule(gotText[1].upper(), beginStudy + datetime.timedelta(days=days_d[gotText[0]]))[25:],
                    getSchedule(gotText[1].upper(),
                                beginStudy + weekDate + datetime.timedelta(days=days_d[gotText[0]]))[25:])
                send_message(event, vk, msg)

            elif re.fullmatch(r"бот погода", event.text.lower()):
                token = "70f6144a2107a75fa96a078fbb4a4660"
                r = requests.get(f"http://api.openweathermap.org/data/2.5/weather?q=moscow&appid={token}&units=metric")
                data = r.json()
                pprint(data)
                send_message(event, vk, "Погода в Москве: " + str(data["weather"][0]["main"]) + "\n" + getWeather(1))

            # Поиск преподавателя
            elif re.match("найти ", event.text.lower()):
                name = event.text.upper()[6] + event.text.lower()[7:]
                professor = professorSearch(event, vk, name)
                if len(professor) == 0:
                    send_message(event, vk, "ХТО ЭТО, Я НЕ ЗНАЮ ХТО ЭТО!!!")

                elif len(professor) == 1:
                    professor_name = professor[0]
                    keyboard_3(event, vk, professor_name)
                else:
                    keyboard_2(event, vk, professor)

            elif event.text in professor:
                professor_name = event.text
                keyboard_3(event, vk, professor_name)

            # Расписание на сегодня
            elif event.text.lower() == "на сегодня" and professor != []:
                if nowDate.weekday() != 6:
                    send_message(event, vk, getProfessor(professor_name, nowDate))
                else:
                    send_message(event, vk, "Пар нет)")

            # Расписание на завтра
            elif event.text.lower() == "на завтра" and professor != []:
                if (nowDate + datetime.timedelta(days=1)).weekday() != 6:
                    send_message(event, vk, getProfessor(professor_name, nowDate + datetime.timedelta(days=1)))
                else:
                    send_message(event, vk, "Завтра выходной")

            # Расписание на данную неделю
            elif event.text.lower() == "на эту неделю" and professor != []:
                weekSchedule_1 = []
                for i in range(6):
                    weekSchedule_1.append(
                        getProfessor(professor_name, nowDate - datetime.timedelta(days=nowDate.weekday() - i)) + "\n\n")
                send_message(event, vk, weekSchedule_1)

            # Расписание на следующую неделю
            elif event.text.lower() == "на следующую неделю" and professor != []:
                weekSchedule_1 = []
                for i in range(6):
                    weekSchedule_1.append(getProfessor(professor_name,
                                                       nowDate + datetime.timedelta(weeks=1) - datetime.timedelta(
                                                           days=nowDate.weekday() - i)) + "\n\n")
                send_message(event, vk, weekSchedule_1)

            # Погода на определенное время
            elif re.fullmatch(r"погода", event.text.lower()):
                keyboard_4(event, vk)

            # Погода сейчас
            elif re.fullmatch(r"сейчас", event.text.lower()):
                attachments = []
                # Добавление изображения
                image = requests.get(getWeather(2), stream=True)
                photo = upload.photo_messages(photos=image.raw)[0]
                attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
                # Отправка вместе с изображением
                vk.messages.send(
                    user_id=event.user_id,
                    attachment=','.join(attachments),
                    random_id=get_random_id(),
                    message="Погода в Москве\n"
                )
                send_message(event, vk, getWeather(1))

            # Погода сегодня
            elif re.fullmatch(r"сегодня", event.text.lower()):
                getPic(1)
                attachments = []
                photo = upload.photo_messages("image.png")[0]
                attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
                vk.messages.send(
                    user_id=event.user_id,
                    attachment=','.join(attachments),
                    random_id=get_random_id(),
                    message="Погода в Москве сегодня\n")
                send_message(event, vk, getWeater_day(1))

            # Погода завтра
            elif re.fullmatch(r"завтра", event.text.lower()):
                getPic(2)
                attachments = []
                photo = upload.photo_messages("image.png")[0]
                attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
                vk.messages.send(
                    user_id=event.user_id,
                    attachment=','.join(attachments),
                    random_id=get_random_id(),
                    message="Погода в Москве завтра\n")
                send_message(event, vk, getWeater_day(2))

            # Погода на 5 дней
            elif re.fullmatch(r"на 5 дней", event.text.lower()):
                day_1 = datetime.datetime.now().date()
                day_2 = day_1 + datetime.timedelta(days=4)
                day_1 = day_1.strftime("%d.%m")
                day_2 = day_2.strftime("%d.%m")
                getPic(3)
                attachments = []
                photo = upload.photo_messages("image.png")[0]
                attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

                vk.messages.send(
                    user_id=event.user_id,
                    attachment=','.join(attachments),
                    random_id=get_random_id(),
                    message=f"Погода в Москве c {day_1} по {day_2}\n")
                send_message(event, vk, getWeater_day5())


            # Корона
            elif re.fullmatch(r"корона", event.text.lower()):
                show_corona_all_stat(event.user_id)
            # Корона и область писать например "Корона Мурманская"
            elif re.fullmatch(r"корона \w{1,30}", event.text.lower()):
                show_corona_region_stat(event, event.text)


if __name__ == '__main__':
    main()
